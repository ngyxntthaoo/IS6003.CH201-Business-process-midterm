"""
report.py — Xuat bao cao Excel dep voi openpyxl
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

TYPE_NAME = {1: "Event", 2: "Task", 3: "Gateway XOR", 4: "Gateway OR"}

C_NAVY    = "1E3A5F"
C_CYAN    = "00B4CC"
C_GREEN   = "16A34A"
C_RED     = "DC2626"
C_PURPLE  = "7C3AED"
C_BLUE_L  = "DBEAFE"
C_GRAY_L  = "F1F5F9"
C_WHITE   = "FFFFFF"
C_DARK    = "1E293B"
C_MID     = "475569"

def _thin():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_font(sz=10): return Font(name="Arial", bold=True,  color=C_WHITE, size=sz)
def _body_font(sz=9): return Font(name="Arial",             color=C_DARK,  size=sz)

def _fmt(v):
    if v is None or v == "" or (isinstance(v, float) and str(v) == "nan"): return "—"
    return str(v)

def _vnd(v):
    try: return f"{int(v):,}đ"
    except: return "—"

def _write_hdr(ws, row, col, value, color=C_NAVY, sz=9):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _hdr_font(sz)
    c.fill      = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _thin()
    return c

def _write_cell(ws, row, col, value, bold=False, bg=C_WHITE,
                align="left", fmt=None, fc=C_DARK):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=fc, size=9)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.border    = _thin()
    c.alignment = Alignment(
        horizontal="center" if align=="center" else ("right" if align=="right" else "left"),
        vertical="center", wrap_text=True)
    if fmt: c.number_format = fmt
    return c

def _section_banner(ws, row, col_start, col_end, title, color=C_NAVY):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=f"  {title}")
    c.font      = Font(name="Arial", bold=True, color=C_WHITE, size=11)
    c.fill      = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = _thin()
    ws.row_dimensions[row].height = 22


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Tóm tắt
# ══════════════════════════════════════════════════════════════════════════════
def _sheet_summary(wb, graph_name, stats, cycles, path_result, now):
    ws = wb.active
    ws.title = "Tóm tắt"
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value     = f"BÁO CÁO PHÂN TÍCH QUY TRÌNH  ·  {graph_name.upper()}"
    t.font      = Font(name="Arial", bold=True, size=16, color=C_WHITE)
    t.fill      = PatternFill("solid", fgColor=C_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 48

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value     = f"Ngày xuất: {now}   |   Hệ thống: Business Process Visualizer"
    s.font      = Font(name="Arial", size=9, color=C_MID, italic=True)
    s.fill      = PatternFill("solid", fgColor=C_BLUE_L)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # KPI cards
    paths    = path_result.get("paths", [])
    min_cost = path_result.get("min_cost", 0)
    max_cost = path_result.get("max_cost", 0)
    kpis = [
        ("Tổng Node",          stats.get("total_nodes", 0),     C_NAVY),
        ("Tổng Edge",          stats.get("total_edges", 0),     C_CYAN),
        ("Task Nodes",         stats.get("total_tasks", 0),     "2563EB"),
        ("Gateway Nodes",      stats.get("total_gateways", 0),  C_PURPLE),
        ("Chu trình",          len(cycles),                     C_RED if cycles else C_GREEN),
        ("Đường đi (Path)",    len(paths),                      C_CYAN),
        ("Chi phí thấp nhất",  min_cost if paths else "—",      C_GREEN),
        ("Chi phí cao nhất",   max_cost if paths else "—",      C_RED),
    ]
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 38
    ws.row_dimensions[6].height = 14
    for i, (lbl, val, color) in enumerate(kpis):
        col = i + 1
        lc = ws.cell(row=4, column=col, value=lbl)
        lc.font      = Font(name="Arial", bold=True, size=8, color=C_WHITE)
        lc.fill      = PatternFill("solid", fgColor=color)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border    = _thin()

        vc = ws.cell(row=5, column=col, value=val)
        vc.font      = Font(name="Arial", bold=True, size=17, color=color)
        vc.fill      = PatternFill("solid", fgColor=C_WHITE)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border    = _thin()
        if "phí" in lbl and isinstance(val, (int, float)) and val:
            vc.number_format = '#,##0"đ"'

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 17

    # Thống kê node
    r = 8
    _section_banner(ws, r, 1, 8, "📊  THỐNG KÊ NODE THEO LOẠI")
    r += 1
    for col, h in enumerate(["Loại Node","Số lượng","Tỉ lệ %","Mô tả"], 1):
        _write_hdr(ws, r, col, h, C_CYAN)
    r += 1
    counts = stats.get("counts", {})
    total  = stats.get("total_nodes", 1) or 1
    descs  = {"Event":"Bắt đầu/Kết thúc","Task":"Bước thực thi",
              "Gateway XOR":"Rẽ nhánh loại trừ","Gateway OR":"Rẽ nhánh hoà nhập"}
    type_colors_bg = {"Event":"E0FFFE","Task":"DBEAFE","Gateway XOR":"FEF3C7","Gateway OR":"F3E8FF"}
    for name, cnt in counts.items():
        bg = type_colors_bg.get(name, C_WHITE)
        _write_cell(ws, r, 1, name, bg=bg)
        _write_cell(ws, r, 2, cnt, bg=bg, align="center")
        pc = _write_cell(ws, r, 3, cnt/total, bg=bg, align="center")
        pc.number_format = "0.0%"
        _write_cell(ws, r, 4, descs.get(name, ""), bg=bg)
        r += 1
    # Total
    _write_cell(ws, r, 1, "TỔNG", bold=True, bg=C_BLUE_L)
    _write_cell(ws, r, 2, f"=SUM(B{r-len(counts)}:B{r-1})", bold=True, bg=C_BLUE_L, align="center")
    _write_cell(ws, r, 3, "100%", bold=True, bg=C_BLUE_L, align="center")
    _write_cell(ws, r, 4, "", bg=C_BLUE_L)

    # Nhận xét
    r += 2
    _section_banner(ws, r, 1, 8, "📝  NHẬN XÉT TỰ ĐỘNG")
    r += 1
    remarks = []
    task_pct = stats.get("total_tasks",0)/total*100
    gw_pct   = stats.get("total_gateways",0)/total*100
    if task_pct >= 50:
        remarks.append(f"✅  Quy trình tập trung thực thi — Task chiếm {task_pct:.0f}% tổng node.")
    if gw_pct >= 30:
        remarks.append(f"⚠️  Nhiều điểm rẽ nhánh — Gateway chiếm {gw_pct:.0f}%, cần kiểm soát luồng.")
    if cycles:
        remarks.append(f"🔴  Phát hiện {len(cycles)} chu trình — nguy cơ vòng lặp vô tận!")
    else:
        remarks.append("🟢  Không có chu trình — quy trình chạy tuyến tính, an toàn.")
    if paths and min_cost < max_cost:
        remarks.append(f"💰  Chênh lệch chi phí giữa các path: {_vnd(max_cost-min_cost)} — ưu tiên path rẻ nhất.")
    for remark in remarks:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws.cell(row=r, column=1, value=remark)
        c.font      = Font(name="Arial", size=9, color=C_DARK)
        c.fill      = PatternFill("solid", fgColor=C_GRAY_L)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _thin()
        ws.row_dimensions[r].height = 18
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Thống kê + Chart
# ══════════════════════════════════════════════════════════════════════════════
def _sheet_stats(wb, graph_name, stats):
    ws = wb.create_sheet("📊 Thống kê")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value     = f"THỐNG KÊ QUY TRÌNH — {graph_name}"
    t.font      = Font(name="Arial", bold=True, size=13, color=C_NAVY)
    t.fill      = PatternFill("solid", fgColor=C_BLUE_L)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    counts  = stats.get("counts", {})
    total   = stats.get("total_nodes", 1) or 1
    descs   = {"Event":"Điểm bắt đầu/kết thúc quy trình","Task":"Bước thực thi công việc cụ thể",
               "Gateway XOR":"Rẽ nhánh loại trừ (1 nhánh)","Gateway OR":"Rẽ nhánh hoà nhập (≥1 nhánh)"}

    r = 3
    for col, h in enumerate(["Loại Node","Số lượng","Tỉ lệ %","Mô tả"], 1):
        _write_hdr(ws, r, col, h, C_NAVY)
    data_start = r + 1
    r += 1

    for name, cnt in counts.items():
        bg = C_GRAY_L if r % 2 == 0 else C_WHITE
        _write_cell(ws, r, 1, name, bg=bg)
        _write_cell(ws, r, 2, cnt,  bg=bg, align="center")
        pc = _write_cell(ws, r, 3, cnt/total, bg=bg, align="center")
        pc.number_format = "0.0%"
        _write_cell(ws, r, 4, descs.get(name,""), bg=bg)
        r += 1
    data_end = r - 1

    _write_cell(ws, r, 1, "TỔNG", bold=True, bg=C_BLUE_L)
    _write_cell(ws, r, 2, f"=SUM(B{data_start}:B{data_end})", bold=True, bg=C_BLUE_L, align="center")
    _write_cell(ws, r, 3, "100%", bold=True, bg=C_BLUE_L, align="center")
    _write_cell(ws, r, 4, "", bg=C_BLUE_L)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 44

    # Bar chart
    bar = BarChart()
    bar.type   = "col"
    bar.title  = "Số lượng Node theo loại"
    bar.y_axis.title = "Số lượng"
    bar.style  = 10; bar.width = 14; bar.height = 10
    bar.add_data(Reference(ws, min_col=2, min_row=data_start-1, max_row=data_end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=data_start, max_row=data_end))
    ws.add_chart(bar, "F3")

    # Pie chart
    pie = PieChart()
    pie.title  = "Tỉ trọng loại Node"
    pie.style  = 10; pie.width = 14; pie.height = 10
    pie.add_data(Reference(ws, min_col=2, min_row=data_start-1, max_row=data_end), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=1, min_row=data_start, max_row=data_end))
    pie.series[0].data_points = [DataPoint(idx=0, explosion=10)]
    ws.add_chart(pie, "F18")


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — Nodes & Edges
# ══════════════════════════════════════════════════════════════════════════════
def _sheet_nodes_edges(wb, graph_data):
    nodes = graph_data["nodes"]
    edges = graph_data["edges"]
    node_label = {n["id"]: n.get("label","") for n in nodes}

    ws = wb.create_sheet("🔵 Nodes & Edges")
    ws.sheet_view.showGridLines = False

    _section_banner(ws, 1, 1, 8, "  DANH SÁCH NODE")
    for col, h in enumerate(["ID","Tên","Loại","Phụ trách","Phòng ban","Nhân sự","Chi phí (đ)","Thiết bị"], 1):
        _write_hdr(ws, 2, col, h, C_NAVY)

    type_bg = {1:"E0FFFE", 2:"DBEAFE", 3:"FEF3C7", 4:"F3E8FF"}
    for i, n in enumerate(nodes):
        r  = i + 3
        bg = type_bg.get(n.get("type",0), C_WHITE)
        _write_cell(ws, r, 1, n["id"],   bg=bg, align="center")
        _write_cell(ws, r, 2, n.get("label",""), bold=True, bg=bg)
        _write_cell(ws, r, 3, TYPE_NAME.get(n.get("type",0),"?"), bg=bg, align="center")
        _write_cell(ws, r, 4, _fmt(n.get("assignee")),   bg=bg)
        _write_cell(ws, r, 5, _fmt(n.get("department")), bg=bg)
        _write_cell(ws, r, 6, int(n.get("headcount") or 0), bg=bg, align="center")
        cc = _write_cell(ws, r, 7, int(n.get("node_cost") or 0), bg=bg, align="right")
        cc.number_format = '#,##0"đ"'
        _write_cell(ws, r, 8, _fmt(n.get("equipment")), bg=bg)

    last_r = 2 + len(nodes)
    ws.merge_cells(start_row=last_r+1, start_column=1, end_row=last_r+1, end_column=6)
    _write_cell(ws, last_r+1, 1, "TỔNG CHI PHÍ", bold=True, bg=C_BLUE_L, align="center")
    tc = _write_cell(ws, last_r+1, 7, f"=SUM(G3:G{last_r})", bold=True, bg=C_BLUE_L, align="right")
    tc.number_format = '#,##0"đ"'
    _write_cell(ws, last_r+1, 8, "", bg=C_BLUE_L)

    for col, w in zip("ABCDEFGH", [6,14,18,18,16,9,16,40]):
        ws.column_dimensions[col].width = w

    # Edge table
    er = last_r + 3
    _section_banner(ws, er, 1, 4, "  DANH SÁCH EDGE")
    for col, h in enumerate(["Từ (ID)","Từ (Tên)","Đến (ID)","Đến (Tên)"], 1):
        _write_hdr(ws, er+1, col, h, C_NAVY)
    for i, (src, tgt) in enumerate(edges):
        r  = er + 2 + i
        bg = C_GRAY_L if i % 2 == 0 else C_WHITE
        _write_cell(ws, r, 1, src, bg=bg, align="center")
        _write_cell(ws, r, 2, node_label.get(src, src), bg=bg)
        _write_cell(ws, r, 3, tgt, bg=bg, align="center")
        _write_cell(ws, r, 4, node_label.get(tgt, tgt), bg=bg)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4 — Chu trình
# ══════════════════════════════════════════════════════════════════════════════
def _sheet_cycles(wb, graph_data, cycles):
    nodes      = graph_data["nodes"]
    node_label = {n["id"]: n.get("label","") for n in nodes}

    ws = wb.create_sheet("🔁 Chu trình")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = "PHÂN TÍCH CHU TRÌNH (CYCLE DETECTION)"
    t.font      = Font(name="Arial", bold=True, size=13, color=C_NAVY)
    t.fill      = PatternFill("solid", fgColor=C_BLUE_L)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    if not cycles:
        ws.merge_cells("A3:F3")
        c = ws["A3"]
        c.value     = "✅  Không phát hiện chu trình nào. Quy trình chạy tuyến tính, an toàn."
        c.font      = Font(name="Arial", bold=True, size=11, color=C_GREEN)
        c.fill      = PatternFill("solid", fgColor="DCFCE7")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _thin()
        ws.row_dimensions[3].height = 28
        return

    ws.merge_cells("A3:F3")
    c = ws["A3"]
    c.value     = f"⚠️  Phát hiện {len(cycles)} chu trình! Kiểm tra điều kiện thoát."
    c.font      = Font(name="Arial", bold=True, size=11, color=C_RED)
    c.fill      = PatternFill("solid", fgColor="FEE2E2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = _thin()
    ws.row_dimensions[3].height = 28

    r = 5
    for col, h in enumerate(["STT","Độ dài","Đường đi","Cảnh báo"], 1):
        _write_hdr(ws, r, col, h, C_RED)
    r += 1
    for i, cycle in enumerate(cycles, 1):
        path_str = " → ".join(node_label.get(nid,"?") for nid in cycle)
        path_str += f" → {node_label.get(cycle[0],'?')}"
        has_gw   = any("Gateway" in TYPE_NAME.get(
            next((n["type"] for n in nodes if n["id"]==nid),0),"") for nid in cycle)
        warning  = "Gateway thiếu điều kiện thoát" if has_gw else "Task loop — thiếu Gateway"
        bg = "FEE2E2" if i % 2 else "FEF2F2"
        _write_cell(ws, r, 1, i,         bg=bg, align="center")
        _write_cell(ws, r, 2, len(cycle), bg=bg, align="center")
        _write_cell(ws, r, 3, path_str,  bg=bg)
        _write_cell(ws, r, 4, warning,   bg=bg, fc=C_RED)
        ws.row_dimensions[r].height = 20
        r += 1

    for col, w in zip("ABCD", [6, 10, 62, 35]):
        ws.column_dimensions[col].width = w


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 5 — Path Analysis
# ══════════════════════════════════════════════════════════════════════════════
def _sheet_paths(wb, path_result):
    paths    = path_result.get("paths", [])
    min_cost = path_result.get("min_cost", 0)
    max_cost = path_result.get("max_cost", 0)

    ws = wb.create_sheet("🗺️ Path Analysis")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value     = "PHÂN TÍCH ĐƯỜNG ĐI (PATH ANALYSIS)"
    t.font      = Font(name="Arial", bold=True, size=13, color=C_NAVY)
    t.fill      = PatternFill("solid", fgColor=C_BLUE_L)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    if not paths:
        ws["A3"] = "Không tìm được đường đi nào."
        return

    # Overview
    r = 3
    _section_banner(ws, r, 1, 4, "  TỔNG QUAN", C_CYAN)
    r += 1
    ov_labels = ["Tổng số Path","Chi phí thấp nhất","Chi phí cao nhất","Tiết kiệm được"]
    ov_values = [len(paths), _vnd(min_cost), _vnd(max_cost), _vnd(max_cost-min_cost)]
    ov_colors = [C_NAVY, C_GREEN, C_RED, "B45309"]
    for col, (lbl, val, color) in enumerate(zip(ov_labels, ov_values, ov_colors), 1):
        _write_hdr(ws, r, col, lbl, color)
        vc = ws.cell(row=r+1, column=col, value=val)
        vc.font      = Font(name="Arial", bold=True, size=13, color=color)
        vc.fill      = PatternFill("solid", fgColor=C_WHITE)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border    = _thin()
        ws.row_dimensions[r+1].height = 28
    r += 3

    # Path table
    _section_banner(ws, r, 1, 8, "  CHI TIẾT TỪNG ĐƯỜNG ĐI", C_NAVY)
    r += 1
    for col, h in enumerate(["Path","Đường đi","Tổng chi phí","Số nhân sự","Phụ trách","Phòng ban","Thiết bị","Ghi chú"], 1):
        _write_hdr(ws, r, col, h, C_NAVY)
    r += 1

    for p in paths:
        is_min = p["cost"] == min_cost and len(paths) > 1
        is_max = p["cost"] == max_cost and len(paths) > 1
        bg     = "DCFCE7" if is_min else ("FEE2E2" if is_max else C_WHITE)
        note   = "🏆 Rẻ nhất" if is_min else ("⚠️ Đắt nhất" if is_max else "")
        cost_fc= C_GREEN if is_min else (C_RED if is_max else C_DARK)

        _write_cell(ws, r, 1, f"Path {p['path_id']}", bold=True, bg=bg, align="center")
        _write_cell(ws, r, 2, p["label"], bg=bg)
        cc = _write_cell(ws, r, 3, p["cost"], bold=True, bg=bg, align="right", fc=cost_fc)
        cc.number_format = '#,##0"đ"'
        _write_cell(ws, r, 4, len(p["people"]), bg=bg, align="center")
        _write_cell(ws, r, 5, "; ".join(x["Phụ trách"] for x in p["people"]) or "—", bg=bg)
        _write_cell(ws, r, 6, "; ".join(dict.fromkeys(x["Phòng ban"] for x in p["people"])) or "—", bg=bg)
        _write_cell(ws, r, 7, "; ".join(p["equipment"]) or "—", bg=bg)
        _write_cell(ws, r, 8, note, bold=True, bg=bg, align="center", fc=cost_fc)
        ws.row_dimensions[r].height = 20
        r += 1

    for col, w in zip("ABCDEFGH", [10, 52, 17, 12, 28, 22, 38, 13]):
        ws.column_dimensions[col].width = w

    # Bar chart chi phí
    r += 2
    chart_start = r
    _write_hdr(ws, r, 1, "Path", C_CYAN)
    _write_hdr(ws, r, 2, "Chi phí (đ)", C_CYAN)
    r += 1
    for p in paths:
        ws.cell(row=r, column=1, value=f"Path {p['path_id']}")
        ws.cell(row=r, column=2, value=p["cost"])
        r += 1

    chart = BarChart()
    chart.type  = "col"
    chart.title = "So sánh chi phí các đường đi"
    chart.y_axis.title = "VNĐ"
    chart.style = 10; chart.width = 18; chart.height = 12
    chart.add_data(Reference(ws, min_col=2, min_row=chart_start, max_row=r-1), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=chart_start+1, max_row=r-1))
    ws.add_chart(chart, f"D{chart_start}")


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 6 — Nhân sự & Thiết bị
# ══════════════════════════════════════════════════════════════════════════════
def _sheet_resources(wb, path_result):
    paths = path_result.get("paths", [])
    ws    = wb.create_sheet("👥 Nhân sự & Thiết bị")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value     = "NHÂN SỰ & THIẾT BỊ THEO ĐƯỜNG ĐI"
    t.font      = Font(name="Arial", bold=True, size=13, color=C_NAVY)
    t.fill      = PatternFill("solid", fgColor=C_BLUE_L)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    if not paths:
        ws["A3"] = "Không có dữ liệu."
        return

    path_bgs = ["DBEAFE","E0F2FE","DCFCE7","FEF9C3","FCE7F3","F3E8FF"]

    r = 3
    _section_banner(ws, r, 1, 4, "  👥  PHÂN BỔ NHÂN SỰ", C_NAVY)
    r += 1
    for col, h in enumerate(["Path","Phụ trách","Phòng ban","Task liên quan"], 1):
        _write_hdr(ws, r, col, h, C_NAVY)
    r += 1
    for p in paths:
        bg = path_bgs[(p["path_id"]-1) % len(path_bgs)]
        for person in p["people"]:
            _write_cell(ws, r, 1, f"Path {p['path_id']}", bold=True, bg=bg, align="center")
            _write_cell(ws, r, 2, person["Phụ trách"], bg=bg)
            _write_cell(ws, r, 3, person["Phòng ban"], bg=bg)
            _write_cell(ws, r, 4, person["Task"],      bg=bg)
            ws.row_dimensions[r].height = 18
            r += 1

    r += 2
    _section_banner(ws, r, 1, 3, "  🔧  DANH MỤC THIẾT BỊ", C_NAVY)
    r += 1
    for col, h in enumerate(["Path","STT","Thiết bị / Công cụ"], 1):
        _write_hdr(ws, r, col, h, C_NAVY)
    r += 1
    for p in paths:
        bg = path_bgs[(p["path_id"]-1) % len(path_bgs)]
        for j, eq in enumerate(p["equipment"], 1):
            _write_cell(ws, r, 1, f"Path {p['path_id']}", bold=True, bg=bg, align="center")
            _write_cell(ws, r, 2, j,  bg=bg, align="center")
            _write_cell(ws, r, 3, eq, bg=bg)
            ws.row_dimensions[r].height = 18
            r += 1

    for col, w in zip("ABCD", [12, 22, 22, 30]):
        ws.column_dimensions[col].width = w


# ══════════════════════════════════════════════════════════════════════════════
# Public API
# ══════════════════════════════════════════════════════════════════════════════
def generate_excel_report(graph_name, graph_data, stats, cycles, path_result) -> bytes:
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    wb  = Workbook()

    _sheet_summary    (wb, graph_name, stats, cycles, path_result, now)
    _sheet_stats      (wb, graph_name, stats)
    _sheet_nodes_edges(wb, graph_data)
    _sheet_cycles     (wb, graph_data, cycles)
    _sheet_paths      (wb, path_result)
    _sheet_resources  (wb, path_result)

    tab_colors = [C_NAVY, "0EA5E9", "6366F1", C_RED, C_GREEN, "A855F7"]
    for ws, color in zip(wb.worksheets, tab_colors):
        ws.sheet_properties.tabColor = color

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_export(graph_name, graph_data, stats, cycles, path_result):
    import streamlit as st

    st.subheader("📤 Xuất báo cáo Excel")
    st.markdown(
        "Báo cáo gồm **6 sheet** định dạng chuyên nghiệp: "
        "**Tóm tắt** · **Thống kê** · **Nodes & Edges** · **Chu trình** · **Path Analysis** · **Nhân sự & Thiết bị**"
    )

    col1, col2 = st.columns([1, 2])
    with col1:
        btn = st.button("🔄 Tạo & Tải báo cáo", type="primary", use_container_width=True)
    with col2:
        st.caption(
            f"📦 Sẽ xuất: **{stats.get('total_nodes',0)} nodes** · "
            f"**{stats.get('total_edges',0)} edges** · "
            f"**{len(cycles)} chu trình** · "
            f"**{len(path_result.get('paths',[]))} đường đi**"
        )

    if btn:
        with st.spinner("Đang tạo báo cáo..."):
            excel_bytes = generate_excel_report(graph_name, graph_data, stats, cycles, path_result)
        filename = (
            "BaoCao_"
            + graph_name.replace(" ","_").replace("—","").replace("/","").strip()
            + "_" + datetime.now().strftime("%Y%m%d_%H%M")
            + ".xlsx"
        )
        st.download_button(
            label="📥 Tải xuống Excel",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.success("✅ Báo cáo đã sẵn sàng!")
